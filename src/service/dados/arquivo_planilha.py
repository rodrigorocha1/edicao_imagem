from typing import Generator, Tuple
from src.service.dados.arquivo import Arquivo
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook
from typing import Tuple


class ArquivoPlanilha(Arquivo):

    def __init__(self, nome_arquivo) -> None:
        super().__init__(nome_arquivo)
        self.__planilha = self._abrir_planilha()
        self.__nome_aba = self.__planilha.active.title
        self.__aba = self.__planilha[self.__nome_aba]

    def _abrir_planilha(self) -> Workbook:

        planilha = load_workbook(self._caminho_arquivo)
        return planilha

    def listar_dados(self) -> Generator[Tuple[str], None, None]:
        for linha in self.__aba.iter_rows(min_row=2, max_col=7):
            nome_curso, nome_participante, tipo_de_participacao, data_inicio, data_termino, carga_horaria, data_emissao_certificado = linha[
                :7]
            yield nome_curso.value, nome_participante.value, tipo_de_participacao.value, data_inicio.value, data_termino.value, carga_horaria.value, data_emissao_certificado.value
